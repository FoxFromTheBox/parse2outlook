import openpyxl
import datetime
from datetime import datetime
import win32com.client
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace('MAPI')
wb = openpyxl.load_workbook(r'C:\Users\RakitinIS\Documents\parser\input\input.xlsx')
sh = wb.active
wb.iso_dates = True

def addevent(start, subject):
    import win32com.client
    oOutlook = win32com.client.Dispatch("Outlook.Application")
    appointment = oOutlook.CreateItem(1) # 1=outlook appointment item
    appointment.Start = start
    appointment.Subject = subject
    appointment.Duration = 60
    appointment.ReminderSet = False
    appointment.Move(sharedCalendar)
    return

for i in range(1, 36):
  if (sh.cell(row = i, column = 1).value):
    subject = sh.cell(row = i, column = 1).value
    Date = sh.cell(row = i, column = 3).value
    Time = sh.cell(row = i, column = 4).value
    DateNoTime = Date.strftime("%Y-%m-%d")
    start = str(DateNoTime) + ' ' + str(Time)
    print(subject)
    print(start)
    print('')
    addevent(start, subject)